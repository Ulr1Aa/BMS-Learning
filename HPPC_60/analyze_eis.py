import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import glob
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def find_eis_directory():
    """自动查找当前目录下的 EIS 数据文件夹。
    
    优先尝试脚本所在目录（兼容在 IDE 中直接打开脚本），
    其次尝试当前工作目录（兼容命令行运行）。
    """
    # 候选目录：脚本所在目录 和 当前工作目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [script_dir]
    if os.getcwd() != script_dir:
        candidates.append(os.getcwd())
    
    for base_dir in candidates:
        eis_dirs = glob.glob(os.path.join(base_dir, "*EIS"))
        eis_dirs = [d for d in eis_dirs if os.path.isdir(d) and not d.endswith("_anay")]
        if eis_dirs:
            eis_dir = eis_dirs[0]
            output_dir = eis_dir + "_anay"
            return eis_dir, output_dir
    
    print("  未在脚本所在目录或当前工作目录下找到 EIS 子目录")
    print(f"  脚本所在目录: {script_dir}")
    print(f"  当前工作目录: {os.getcwd()}")
    return None, None


# 设置输出目录
EIS_DIR, OUTPUT_DIR = find_eis_directory()
if EIS_DIR is None:
    raise RuntimeError(
        "未能自动找到 EIS 数据目录。请确保脚本放在正确的 SOC 目录下，"
        "且该目录中包含形如 'HPPC_XX%_YY℃_EIS' 的文件夹。"
    )

os.makedirs(OUTPUT_DIR, exist_ok=True)

# 设置中文字体（Windows）
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


def read_eis_csv(fpath):
    df = pd.read_csv(fpath, encoding='gbk', skiprows=31)
    return df


def auto_label_from_filename(fname):
    name = os.path.splitext(fname)[0]
    
    # 模式1: 两位数字 + STEADY (如 75STEADY, 95STEADY)
    m = re.match(r'^(\d{2})STEADY$', name, re.IGNORECASE)
    if m:
        return f"{m.group(1)}%SOC Steady"
    
    # 模式2: 两位数字 + 一位数字 + C (如 751C → 75%SOC 1C)
    m = re.match(r'^(\d{2})(\d)C$', name, re.IGNORECASE)
    if m:
        return f"{m.group(1)}%SOC {m.group(2)}C"
    
    # 模式3: 纯 STEADY
    if re.match(r'^STEADY$', name, re.IGNORECASE):
        return "Steady"
    
    # 模式4: 纯数字 + C (如 1C, 2C)
    m = re.match(r'^(\d)C$', name, re.IGNORECASE)
    if m:
        return f"{m.group(1)}C"
    
    # 模式5: 两位数字 + 任意后缀
    m = re.match(r'^(\d{2})(.+)$', name)
    if m:
        return f"{m.group(1)}%SOC {m.group(2)}"
    
    # 默认：直接返回文件名
    return name


def discover_eis_files(directory):
    pattern = os.path.join(directory, "*.csv")
    csv_files = sorted(glob.glob(pattern))
    eis_files = []
    for fpath in csv_files:
        fname = os.path.basename(fpath)
        label = auto_label_from_filename(fname)
        eis_files.append((fname, label))
    return eis_files


def extract_soc(eis_dir):
    """从目录路径自动提取 SOC，例如 HPPC_95%_25C_EIS → 95%"""
    basename = os.path.basename(eis_dir)
    m = re.search(r'(\d+)%', basename)
    if m:
        return f"{m.group(1)}%"
    parent = os.path.basename(os.path.dirname(eis_dir))
    m = re.search(r'(\d+)', parent)
    if m:
        return f"{m.group(1)}%"
    return "Unknown"


def find_ellipse_region(Re, Im_neg, cross_idx):
    """
    自动识别椭圆（半圆）区域，返回 (apex_idx, ellipse_end_idx, Re_max)
    
    算法：
    1. 对 -Im 进行 5 点移动平均平滑，抑制噪声
    2. 在穿越点之后找到第一个局部极大值（椭圆顶点 apex）
    3. 在顶点之后找到最后一个局部极小值（椭圆底部结束点）
    4. 椭圆区域 = [cross_idx, ellipse_end_idx]
    5. Re_max = 椭圆区域内 Re 的最大值
    """
    n = len(Im_neg)
    window = 5
    if n >= window:
        smoothed = np.convolve(Im_neg, np.ones(window)/window, mode='same')
        half = window // 2
        for i in range(half):
            smoothed[i] = np.mean(Im_neg[:i+half+1])
        for i in range(n - half, n):
            smoothed[i] = np.mean(Im_neg[i-half:])
    else:
        smoothed = Im_neg.copy()

    search_end = min(n, int(n * 0.6))
    apex_idx = cross_idx
    for i in range(cross_idx + 1, search_end - 1):
        if smoothed[i] > smoothed[i-1] and smoothed[i] > smoothed[i+1]:
            apex_idx = i
            break
    if apex_idx == cross_idx:
        segment = smoothed[cross_idx:search_end]
        if len(segment) > 0:
            apex_idx = cross_idx + np.argmax(segment)

    ellipse_end_idx = apex_idx
    last_min_idx = apex_idx
    for i in range(apex_idx + 1, n - 1):
        if smoothed[i] < smoothed[i-1] and smoothed[i] < smoothed[i+1]:
            last_min_idx = i
    if last_min_idx == apex_idx:
        for i in range(apex_idx + 2, n - 1):
            if smoothed[i] > smoothed[i-1] and smoothed[i-1] > smoothed[i-2]:
                last_min_idx = i - 1
                break
    ellipse_end_idx = last_min_idx

    ellipse_Re = Re[cross_idx:ellipse_end_idx+1]
    Re_max = float(np.max(ellipse_Re)) if len(ellipse_Re) > 0 else Re[-1]
    return apex_idx, ellipse_end_idx, Re_max


def calculate_diffusion_resistance(Re, Im_neg, freq, ellipse_end_idx, R_ohm, R_ct):
    """
    用 Warburg 公式独立计算扩散阻抗。
    
    算法：
    1. 扩散区：Re_w = Re - R_ohm - R_ct > 0 的点
    2. Warburg 关系：Re_w = sigma / sqrt(omega)，其中 omega = 2*pi*f
    3. 对 Re_w 与 1/sqrt(omega) 进行过原点线性拟合，斜率 = sigma
    4. 在最低频率处：R_diff = sigma / sqrt(omega_min)
    5. 扩散阻抗模值：|Z_w| = sqrt(2) * R_diff
    
    如果扩散区有效点不足 2 个，回退到简单差值法。
    """
    R_total = Re[-1]
    omega = 2 * np.pi * freq
    Re_w = Re - R_ohm - R_ct
    mask = (np.arange(len(Re)) > ellipse_end_idx) & (Re_w > 0)
    
    if mask.sum() < 2:
        R_diff = R_total - R_ohm - R_ct
        return R_diff, 0.0, 0.0, R_diff, 0.0
    
    f_diff = freq[mask]
    Re_w_diff = Re_w[mask]
    omega_diff = omega[mask]
    inv_sqrt_omega = 1.0 / np.sqrt(omega_diff)
    
    sigma = np.sum(Re_w_diff * inv_sqrt_omega) / np.sum(inv_sqrt_omega ** 2)
    
    y_pred = sigma * inv_sqrt_omega
    ss_res = np.sum((Re_w_diff - y_pred) ** 2)
    ss_tot = np.sum((Re_w_diff - np.mean(Re_w_diff)) ** 2)
    r_squared_warburg = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
    
    omega_min = omega[-1]
    R_diff = sigma / np.sqrt(omega_min)
    Z_w_mag = np.sqrt(2) * R_diff
    R_diff_simple = R_total - R_ohm - R_ct
    
    return R_diff, sigma, Z_w_mag, R_diff_simple, r_squared_warburg


def analyze_eis():
    eis_files = discover_eis_files(EIS_DIR)
    if not eis_files:
        print(f"  在 {EIS_DIR} 中没有找到任何 CSV 文件")
        return {}, pd.DataFrame()

    soc = extract_soc(EIS_DIR)
    print(f"  检测到 SOC: {soc}")
    print(f"  EIS 目录: {EIS_DIR}")
    print(f"  输出目录: {OUTPUT_DIR}")

    eis_data = {}
    for fname, label in eis_files:
        fpath = os.path.join(EIS_DIR, fname)
        try:
            df = read_eis_csv(fpath)
        except Exception as e:
            print(f"  读取 {fname} 失败: {e}")
            continue
        eis_data[label] = df
        print(f"\n=== {label} ===")
        print(f"  数据点数: {len(df)}")
        print(f"  频率范围: {df['Frequency(Hz)'].iloc[0]:.2f} ~ {df['Frequency(Hz)'].iloc[-1]:.4f} Hz")
        print(f"  电压: {df['IC2(VM)'].iloc[0]:.1f} mV, 温度: {df['IC2(TM)'].iloc[0]:.1f} °C")

    if not eis_data:
        print("  没有成功读取任何 EIS 数据")
        return {}, pd.DataFrame()

    fig, ax = plt.subplots(figsize=(10, 8), dpi=120)
    color_cycle = ['#1a476f', '#c41e3a', '#2d8f47', '#7a3e9d',
                   '#d47500', '#0096a6', '#c15c17', '#5a9b5a']
    marker_cycle = ['o', 's', '^', 'D', 'v', '<', '>', 'p']

    results = []

    for idx, (label, df) in enumerate(eis_data.items()):
        Re = df['IC2(Re)'].values
        Im = df['IC2(-Im)'].values
        freq = df['Frequency(Hz)'].values

        color = color_cycle[idx % len(color_cycle)]
        marker = marker_cycle[idx % len(marker_cycle)]

        ax.plot(Re, Im, c=color, alpha=0.6, linewidth=2, zorder=2)
        ax.scatter(Re, Im, c=color, s=50, marker=marker,
                   alpha=0.9, edgecolors='white', linewidth=1.2, zorder=3, label=label)

        # 1) 计算 R_ohm：-Im 从负到正穿越实轴的位置
        R_ohm = Re[0]
        cross_idx = 0
        for i in range(len(Im) - 1):
            if Im[i] < 0 and Im[i+1] >= 0:
                t = abs(Im[i]) / (abs(Im[i]) + abs(Im[i+1]))
                R_ohm = Re[i] + t * (Re[i+1] - Re[i])
                cross_idx = i
                break

        # 2) 识别椭圆区域，计算 R_ct
        apex_idx, ellipse_end_idx, Re_max = find_ellipse_region(Re, Im, cross_idx)
        R_ct = Re_max - R_ohm

        # 3) 计算 R_diff（扩散阻抗）—— Warburg 拟合
        R_diff, sigma, Z_w_mag, R_diff_simple, r_sq_warburg = calculate_diffusion_resistance(
            Re, Im, freq, ellipse_end_idx, R_ohm, R_ct)
        R_total = Re[-1]

        # 4) 在图上标记关键点
        # 高频起点（R_ohm）- 空心圆
        ax.scatter(Re[cross_idx], Im[cross_idx], s=120, marker='o',
                   facecolors='white', edgecolors='black', linewidth=2, zorder=5)
        # 椭圆顶点（apex）- 橙色三角
        ax.scatter(Re[apex_idx], Im[apex_idx], c='orange', s=100, marker='^',
                   edgecolors='black', linewidth=1.5, zorder=6)
        # 椭圆结束点 - 紫色倒三角
        ax.scatter(Re[ellipse_end_idx], Im[ellipse_end_idx], c='purple', s=100, marker='v',
                   edgecolors='black', linewidth=1.5, zorder=6)
        # 低频终点（R_total）- 实心圆
        ax.scatter(Re[-1], Im[-1], c=color, s=120, marker='o',
                   edgecolors='black', linewidth=2, zorder=5)

        # 绘制 Warburg 扩散拟合线
        omega = 2 * np.pi * freq
        if ellipse_end_idx < len(Re) - 1:
            omega_fit = np.linspace(omega[ellipse_end_idx], omega[-1] * 0.5, 50)
            Re_fit = R_ohm + R_ct + sigma / np.sqrt(omega_fit)
            Im_fit = -sigma / np.sqrt(omega_fit)
            valid = (Re_fit >= Re[ellipse_end_idx]) & (Re_fit <= R_total + 50)
            if valid.sum() > 1:
                ax.plot(Re_fit[valid], Im_fit[valid], '--', color=color, alpha=0.5, linewidth=1.5, zorder=1)

        # 添加频率方向箭头（从高频到低频）
        if len(Re) > 5:
            arrow_idx = len(Re) // 3
            ax.annotate('', xy=(Re[arrow_idx], Im[arrow_idx]),
                        xytext=(Re[arrow_idx-2], Im[arrow_idx-2]),
                        arrowprops=dict(arrowstyle='->', color=color,
                                        linewidth=2, alpha=0.7))

        # 汇总结果
        results.append({
            'Condition': label,
            'R_ohm_mΩ': round(R_ohm, 2),
            'R_ct_mΩ': round(R_ct, 2),
            'R_diff_mΩ': round(R_diff, 2),
            'R_total_mΩ': round(R_total, 2),
            'Sigma_Warburg': round(sigma, 2),
            '|Z_w|_mΩ': round(Z_w_mag, 2),
            'R_diff_simple_mΩ': round(R_diff_simple, 2),
            'Warburg_R2': round(r_sq_warburg, 3),
            'Ellipse_Re_max_mΩ': round(Re_max, 2),
            'Apex_Freq_Hz': round(freq[apex_idx], 4),
            'EllipseEnd_Freq_Hz': round(freq[ellipse_end_idx], 4),
            '|Z|_max_mΩ': round(df['IC2(|Z|)'].max(), 2),
            '|Z|_min_mΩ': round(df['IC2(|Z|)'].min(), 2),
            'Max_-Im_mΩ': round(Im.max(), 2)
        })

    # 坐标轴设置
    ax.set_xlabel('Re(Z) [mΩ]', fontsize=14, fontweight='bold', labelpad=12)
    ax.set_ylabel('-Im(Z) [mΩ]', fontsize=14, fontweight='bold', labelpad=12)
    ax.set_title('Nyquist Plot', fontsize=16, fontweight='bold', pad=20)
    ax.grid(True, which='major', linestyle='-', linewidth=0.8, alpha=0.3, color='gray')
    ax.grid(True, which='minor', linestyle=':', linewidth=0.5, alpha=0.2, color='gray')
    ax.minorticks_on()
    ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize=11,
              framealpha=0.95, edgecolor='gray',
              borderpad=1.2, labelspacing=1.0, shadow=True)
    ax.set_aspect('equal', adjustable='datalim')
    ax.spines['top'].set_linewidth(1.2)
    ax.spines['right'].set_linewidth(1.2)
    ax.spines['bottom'].set_linewidth(1.2)
    ax.spines['left'].set_linewidth(1.2)
    ax.tick_params(axis='both', which='major', labelsize=12, width=1.2, length=6)
    ax.tick_params(axis='both', which='minor', width=0.8, length=3)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, 'EIS_Nyquist_Plot.png'), dpi=300, bbox_inches='tight')
    plt.show()

    # 打印并保存 CSV 结果
    results_df = pd.DataFrame(results)
    print("\n========== EIS 阻抗分析结果 ==========")
    print(results_df.to_string(index=False))

    csv_path = os.path.join(OUTPUT_DIR, 'EIS_Impedance_Results.csv')
    results_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"\nCSV 结果已保存: {csv_path}")

    # 保存 Excel 结果（按用户要求的格式）
    excel_path = os.path.join(OUTPUT_DIR, 'EIS_Impedance_Results.xlsx')
    save_excel(results, soc, excel_path)
    print(f"Excel 结果已保存: {excel_path}")

    return eis_data, results_df


def save_excel(results, soc, excel_path):
    """按用户要求的格式保存为 Excel：SOC 合并单元格，列：电流 / R_ohm / R_ct / R_diff / R_total"""
    wb = Workbook()
    ws = wb.active
    ws.title = "EIS 阻抗汇总"
    
    # 表头
    headers = ['电状态(SOC)', '电流', '欧姆内阻(RΩ)', '电荷转移阻抗Rct', '扩散阻抗RΩ', '总阻抗']
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # 写入数据行
    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=soc)
        ws.cell(row=row_idx, column=2, value=r['Condition'])
        ws.cell(row=row_idx, column=3, value=r['R_ohm_mΩ'])
        ws.cell(row=row_idx, column=4, value=r['R_ct_mΩ'])
        ws.cell(row=row_idx, column=5, value=r['R_diff_mΩ'])
        ws.cell(row=row_idx, column=6, value=r['R_total_mΩ'])
    
    # 合并 SOC 列（如果有多行）
    if len(results) > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=1+len(results), end_column=1)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 统一设置数据区域样式
    for row in range(2, 2 + len(results)):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if col >= 3:
                cell.number_format = '0.00'
    
    # 调整列宽
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(excel_path)


if __name__ == '__main__':
    print("=" * 60)
    print("EIS 数据分析脚本")
    print("=" * 60)
    analyze_eis()
